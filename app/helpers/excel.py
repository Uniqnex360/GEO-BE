from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


class ExcelTemplateBulider:
    """
    reusable class to create template excel files for bulk upload
    """

    def __init__(
        self,
        headers: dict,
        sheet_name="template",
        data: list[dict] | None = None,
        example: bool = False,
    ):
        self.headers = headers
        self.sheet_name = sheet_name
        self.data = data or []
        self.example = example

        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = sheet_name

        self._init_styles()

    def _init_styles(self):
        """
        common styles for the workbook
        1. fonts
        2. colors
        3. headers
        """
        # 1. fonts
        self.header_font = Font(bold=True)
        self.placeholder_font = Font(color="888888", italic=True)

        # 2. colors
        """
        color styles
        Required field     → FFF2CC (light yellow)
        Error highlight    → FFC7CE (light red)
        Success/valid data → C6EFCE (light green)
        Info column        → D9E1F2 (light blue)
        Disabled cells     → F2F2F2 (light gray)
        """
        self.required_color = PatternFill(
            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
        )
        self.error_color = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )
        self.success_color = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )
        self.info_color = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )
        self.disabled_color = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
        )

        # 3. headers
        self.header_alignment = Alignment(vertical="center")

    def build(self):
        """
        flow:
        -> create headers
        -> add data
        -> fix column width
        -> freeze the header # css sticky
        """

        self._create_headers()
        self._add_data()
        self._apply_dropdowns()
        self._auto_fit_columns()
        self._freeze_header()

        return self.wb

    def _create_headers(self):
        """creates headers from schema"""

        for i, header in enumerate(self.headers, start=1):
            # creating cell
            cell = self.ws.cell(
                row=1, column=i
            )  # for headers row always be 1, i-> index
            cell.value = header["identity"]
            cell.font = self.header_font
            cell.alignment = self.header_alignment

            # handle required fields
            if header.get("required"):
                cell.fill = self.required_color

            elif header.get("error"):
                cell.fill = self.error_color

            elif header.get("disabled"):
                cell.fill = self.disabled_color

            # handling comments
            if header.get("comment"):
                cell.comment = Comment(header["comment"], "System")

    def _add_data(self):
        """adds the data into rows
        if example is set to true the data becomes like a place holder in a search box
        """

        row_count = 2

        for row in self.data:

            for i, header in enumerate(self.headers, start=1):

                value = row.get(header["id"])

                if value is not None:

                    cell = self.ws.cell(row=row_count, column=i)
                    cell.value = value

                    if self.example:
                        cell.font = self.placeholder_font

            row_count += 1

    def _apply_dropdowns(self):
        """applyes dropdowns to headers only if example is set to True"""

        # NOTE: for more professional way use dropdown values in Meta sheet

        if not self.example:
            return

        for i, header in enumerate(self.headers, start=1):

            dropdown = header.get("dropdown")

            if not dropdown:
                continue

            values = dropdown.get("values", [])
            allow_blank = dropdown.get("allow_blank", True)
            allow_creation = dropdown.get("allow_creation", False)

            col_letter = get_column_letter(i)  # returns excel col Ex: 1 -> A
            dv = DataValidation(
                type="list",
                formula1=f'"{",".join(values)}"',  # excel will accept this : EX: '"OptionA,OptionB"'
                allow_blank=allow_blank,
                showDropDown=False,
            )
            # user can create new values that is not present in the Options
            if allow_creation:
                dv.showErrorMessage = False

            dv.promptTitle = header["identity"]
            dv.prompt = header.get("comment", "")

            self.ws.add_data_validation(dv)

            # for now i have set add validation for first 100 rows increase as per your needs
            dv.add(f"{col_letter}2:{col_letter}100")

    def _auto_fit_columns(self):
        """gives proper space for headers"""

        for i, header in enumerate(self.headers, start=1):
            column_letter = get_column_letter(
                i
            )  # returns excel equivalent col Ex: 1 -> A, 2 -> B

            # if header is large then use it size else use commant
            # max width allowed 50
            header_lenght = len(header["identity"])
            comment_lenght = len(header.get("comment", ""))

            width = max(header_lenght, comment_lenght)
            width = min(width + 5, 35)

            self.ws.column_dimensions[column_letter].width = width

    def _freeze_header(self):
        """freezes header of the file while scroll [sticky equivant in css]
        for better user experiance
        """

        self.ws.freeze_panes = "A2"


async def validate_headers(headers: list, template_headers: list[dict]) -> list:
    """Validates excel file header strings against configuration keys.

    Extracts explicit field names out of standard definition maps and
    determines what attributes are entirely missing from row 0.
    """

    # 1. Standardize template IDs (lowercase and stripped)
    target_ids = [str(item["id"]).strip().lower() for item in template_headers]

    # 2. NEW: Standardize incoming Excel headers (lowercase, stripped, and spaces to underscores)
    cleaned_headers = [str(h).strip().lower().replace(" ", "_") for h in headers]
    print('log', target_ids, cleaned_headers)
    # 3. Compare the cleaned headers
    missing_fields = [target for target in target_ids if target not in cleaned_headers]

    return missing_fields
