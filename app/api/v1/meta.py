from io import BytesIO
from pydantic import BaseModel
from typing import Optional, List, Dict, Any
from openpyxl import load_workbook
from sqlalchemy import select, func, desc, or_, asc
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, Query


from app.core.database import get_db
from app.core.security import validate_jwt_token
from app.core.permission import require_super_admin
from app.models import MetaTable


class CategoryListRequest(BaseModel):
    search: str = ""
    offset: int = 0
    limit: int = 50


router = APIRouter()

BATCH_SIZE = 500


@router.post("/bulk-upload/")
async def bulk_upload_meta_data(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_super_admin),
):
    try:
        contents = await file.read()
        workbook = load_workbook(BytesIO(contents), data_only=True)
        sheet = workbook.active

        rows = list(sheet.iter_rows(min_row=2, values_only=True))  # Skip header

        if not rows:
            raise HTTPException(status_code=400, detail="Excel file is empty")

        objects = []

        for row in rows:
            if not any(row):
                continue

            objects.append(
                MetaTable(
                    industry_name=row[0],
                    taxonomy=row[1],
                    category_name=row[2],
                )
            )

        inserted = 0
        skipped = 0

        for i in range(0, len(objects), BATCH_SIZE):
            batch = objects[i : i + BATCH_SIZE]

            try:
                db.add_all(batch)
                await db.commit()
                inserted += len(batch)

            except IntegrityError:
                await db.rollback()

                # Insert one-by-one so duplicate rows are skipped
                for obj in batch:
                    try:
                        db.add(obj)
                        await db.commit()
                        inserted += 1
                    except IntegrityError:
                        await db.rollback()
                        skipped += 1

        return {
            "message": "Bulk upload completed.",
            "inserted": inserted,
            "skipped_duplicates": skipped,
            "total_rows": len(objects),
        }

    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/category/list")
async def get_category_meta_list(
    search: str = Query(None, description="Search category"),
    offset: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(validate_jwt_token),
):
    stmt = select(MetaTable.category_name).distinct()

    if search:
        stmt = stmt.where(MetaTable.category_name.ilike(f"%{search}%"))

    stmt = stmt.order_by(MetaTable.category_name).offset(offset).limit(limit + 1)

    result = await db.execute(stmt)
    categories = result.scalars().all()

    has_more = len(categories) > limit

    if has_more:
        categories.pop()

    return {
        "items": [
            {
                "id": category,
                "identity": category,
            }
            for category in categories
        ],
        "has_more": has_more,
        "next_offset": offset + limit if has_more else None,
    }


@router.get("/industry/list")
async def get_industry_meta_list(
    search: str = Query(None, description="Search category"),
    offset: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(validate_jwt_token),
):
    stmt = select(MetaTable.industry_name).distinct()

    if search:
        stmt = stmt.where(MetaTable.industry_name.ilike(f"%{search}%"))

    stmt = stmt.order_by(MetaTable.industry_name).offset(offset).limit(limit + 1)

    result = await db.execute(stmt)
    industries = result.scalars().all()

    has_more = len(industries) > limit

    if has_more:
        industries.pop()

    return {
        "items": [
            {
                "id": industry,
                "identity": industry,
            }
            for industry in industries
        ],
        "has_more": has_more,
        "next_offset": offset + limit if has_more else None,
    }


@router.get("/taxonomy/list")
async def get_taxonomy_list(
    search: Optional[str] = Query(None),
    page: Optional[int] = Query(1, ge=1),
    limit: Optional[int] = Query(24, ge=1),
    sort_by: Optional[str] = Query(
        "created_at",
        description="Field to sort by: 'category_name', 'industry_name', 'end_category', or 'created_at'",
    ),
    sort_order: Optional[str] = Query(
        "desc", regex="^(asc|desc)$", description="Sort direction: 'asc' or 'desc'"
    ),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(validate_jwt_token),
):
    """
    List taxonomies/meta entries with multi-field search, end_category extraction,
    dynamic sorting, and standardized pagination.
    """
    current_page = page if page is not None else 1
    current_limit = limit if limit is not None else 24

    total_count_col = func.count().over().label("total_count_val")
    stmt = select(MetaTable, total_count_col)

    # 1. Multi-field search filter
    search_filter = None
    if search and search.strip():
        search_pattern = f"%{search.strip()}%"
        search_filter = or_(
            MetaTable.category_name.ilike(search_pattern),
            MetaTable.industry_name.ilike(search_pattern),
            MetaTable.taxonomy.ilike(search_pattern),
        )
        stmt = stmt.where(search_filter)

    # 2. Build the SQL expression for end_category (PostgreSQL example)
    # Extracts the last string segment after '>' and trims whitespace
    db_end_category = func.trim(
        func.split_part(
            func.coalesce(MetaTable.taxonomy, MetaTable.category_name), ">", -1
        )
    )

    # 3. Dynamic Sorting Resolution
    sort_column_map = {
        "category_name": MetaTable.category_name,
        "industry_name": MetaTable.industry_name,
        "end_category": db_end_category,
        "created_at": MetaTable.created_at,
    }

    # Fall back to created_at if an invalid field name is passed
    target_column = sort_column_map.get(sort_by, MetaTable.created_at)

    direction_func = desc if sort_order == "asc" else desc  # Default check
    if sort_order and sort_order.lower() == "asc":
        direction_func = asc
    else:
        direction_func = desc

    # 4. Apply ordering and pagination
    stmt = (
        stmt.order_by(direction_func(target_column))
        .offset((current_page - 1) * current_limit)
        .limit(current_limit)
    )

    result = await db.execute(stmt)
    rows = result.all()

    # 5. Extract items and window count
    items: List[Dict[str, Any]] = []
    total = rows[0].total_count_val if rows else 0

    for row in rows:
        meta_obj: MetaTable = row[0]

        end_cat = None
        if meta_obj.taxonomy:
            end_cat = meta_obj.taxonomy.split(">")[-1].strip()
        elif meta_obj.category_name:
            end_cat = meta_obj.category_name

        item_dict = {
            "id": meta_obj.id,
            "category_name": meta_obj.category_name,
            "industry_name": meta_obj.industry_name,
            "taxonomy": meta_obj.taxonomy,
            "end_category": end_cat,
            "is_active": getattr(meta_obj, "is_active", True),
            "is_deleted": getattr(meta_obj, "is_deleted", False),
            "created_at": (
                meta_obj.created_at.isoformat() if meta_obj.created_at else None
            ),
            "updated_at": (
                meta_obj.updated_at.isoformat()
                if getattr(meta_obj, "updated_at", None)
                else None
            ),
        }
        items.append(item_dict)

    # 6. Fallback count check if page requested is out of bounds
    if not rows:
        count_stmt = select(func.count(MetaTable.id))
        if search_filter is not None:
            count_stmt = count_stmt.where(search_filter)
        total_res = await db.execute(count_stmt)
        total = total_res.scalar() or 0

    # 7. Response payload
    return {
        "items": items,
        "pagination": {
            "page": current_page,
            "limit": current_limit,
            "total": total,
        },
    }
