from typing import Optional
from io import BytesIO
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    Query,
    BackgroundTasks,
    File,
    UploadFile,
)
from fastapi.responses import StreamingResponse

from app.core.database import get_db, SessionLocal
from app.core.security import validate_jwt_token
from app.helpers import ExcelTemplateBulider, validate_headers
from app.schemas import (
    ProductCreate,
    ProductUpdate,
)
from app.models import Product, Brand
from app.services import ProductService

router = APIRouter()


@router.post("/create/")
async def create_product(
    payload: ProductCreate,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(validate_jwt_token),
):
    """create product"""

    try:
        data = payload.model_dump()

        is_super_admin = user.get("is_super_admin")
        tenant_id = data.get("tenant_id") if is_super_admin else user.get("tenant_id")

        response = await ProductService.create_product(
            db=db, data=payload.model_dump(), user=user, tenant_id=tenant_id
        )

        return {"message": "Product created successfully", "data": response}

    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.put("/update/{product_id}/")
async def update_product(
    product_id: int,
    payload: ProductUpdate,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(validate_jwt_token),
):
    """update product"""

    try:

        response = await ProductService.update_product(
            db=db,
            product_id=product_id,
            data=payload.model_dump(exclude_unset=True),
            user=user,
        )

        return {"message": "Product updated successfully", "data": response}

    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.patch("/status/{product_id}")
async def active_inactive_product(
    product_id: int,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(validate_jwt_token),
):
    """activate/deactivate Product"""

    try:

        response = await ProductService.soft_delete_product(
            db=db, product_id=product_id, user=user
        )

        return {"message": "Product status updated successfully", "data": response}

    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/delete/{product_id}")
async def delete_product(
    product_id: int,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(validate_jwt_token),
):
    """logical delete"""

    try:

        await ProductService.delete_product(db=db, product_id=product_id, user=user)

        return {"message": "Product deleted successfully"}

    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/list/")
async def list_products(
    page: int = Query(1, ge=1),
    limit: int = Query(24, ge=1),
    search: Optional[str] = None,
    sort_by: str = "created_at",
    sort_order: str = "desc",
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(validate_jwt_token),
    tenant_id: Optional[int] = Query(
        None, description="Super admins can filter by a specific tenant ID"
    ),
    brand: Optional[str] = Query(None, description="Comma-separated list of brands"),
):
    """list products"""

    is_super_admin = user.get("is_super_admin", False)
    active_tenant_id = tenant_id if is_super_admin else user.get("tenant_id")

    if sort_by == "analytics.visibility_rate":
        sort_by = "visibility"
    elif sort_by == "analytics.by_engine.chatgpt.visibility_rate":
        sort_by = "visibility_gpt"
    elif sort_by == "analytics.by_engine.gemini.visibility_rate":
        sort_by = "visibility_gemini"
    elif sort_by == "analytics.by_engine.anthropic.visibility_rate":
        sort_by = "visibility_claude"

    try:

        products, total, tenant_states, product_ids = await ProductService.list_products(
            db=db,
            user=user,
            tenant_id=active_tenant_id,
            page=page,
            limit=limit,
            search=search,
            brand=brand,
            sort_by=sort_by,
            sort_order=sort_order,
        )

        return {
            "data": products,
            "tenant_states": tenant_states,
            "pagination": {"page": page, "limit": limit, "total": total},
            "product_ids": product_ids
        }

    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/detail/{product_id}")
async def product_detail(
    product_id: int,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(validate_jwt_token),
):
    """product detail"""

    try:

        data = await ProductService.detail(
            db=db, product_id=product_id, tenant_id=user.get("tenant_id"), user=user
        )

        return data

    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/detail/v2/{product_id}")
async def product_detail(
    product_id: int,
    tab: str = Query(
        "visibility", description="Target specific dashboard tab dataset dynamically"
    ),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(validate_jwt_token),
):
    """Product detail version 2 - Dynamically filtered by active tab state"""
    try:
        data = await ProductService.product_detail_v2(
            db=db,
            product_id=product_id,
            tenant_id=user.get("tenant_id"),
            user=user,
            tab=tab.lower().strip(),
        )
        return data
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


PRODUCT_TEMPLATE_HEADERS = [
    {
        "id": "product_name",
        "identity": "Product Name",
        "required": True,
        "comment": "Name of the Product",
    },
    {
        "id": "brand_name",
        "identity": "Brand Name",
        "required": True,
        "comment": "Name of the Brand",
    },
    {
        "id": "sku",
        "identity": "SKU",
        "required": True,
        "comment": "SKU for the product",
    },
    {
        "id": "mpn",
        "identity": "MPN",
        "required": True,
        "comment": "MPN of the product",
    },
    {
        "id": "product_url",
        "identity": "Product URL",
        "required": True,
        "comment": "URL of the product on the website",
    },
]


PRODUCT_EXAMPLE_DATA = [
    {
        "product_name": "iPhone 16 Pro",
        "brand_name": "Apple",
        "sku": "APL-IP16P-256-BLK",
        "mpn": "MYN03LL/A",
        "product_url": "https://www.example.com/products/iphone-16-pro",
    }
]


@router.get("/bulk-upload-template/")
async def generate_product_template(user: dict = Depends(validate_jwt_token)):
    """api endpoint that streams bulk upload template for chat"""

    builder = ExcelTemplateBulider(
        headers=PRODUCT_TEMPLATE_HEADERS,
        sheet_name="product_template",
        data=PRODUCT_EXAMPLE_DATA,
        example=True,
    )

    wb = builder.build()

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachement; filename=product_template.xlsx"},
    )


async def process_product_row(
    row_data: dict, tenant_id: str | int, user_id: str | int, session_factory
):
    """Worker task executed in the background for each row."""
    async with session_factory() as db:
        try:
            print("background task", row_data, flush=True)

            # 1. Clean and normalize input values safely
            product_name = (
                str(row_data["product_name"]).strip()
                if row_data.get("product_name")
                else None
            )

            if not product_name:
                print(f"[SKIP] Missing product_name for row: {row_data}")
                return

            # Extract raw string value safely
            raw_brand = row_data.get("brand_name")
            brand_str = str(raw_brand).strip() if raw_brand else "Generic"
            print("brand name", brand_str)

            sku = str(row_data["sku"]).strip() if row_data.get("sku") else None
            mpn = str(row_data["mpn"]).strip() if row_data.get("mpn") else None

            # Fixed key check: changed 'product url' to 'product_url'
            product_url = (
                str(row_data["product_url"]).strip()
                if row_data.get("product_url")
                else None
            )

            tenant_id_int = int(tenant_id) if tenant_id is not None else None
            user_id_int = int(user_id) if user_id is not None else None

            # 2. Handle Brand (Find or Create scoped to tenant_id)
            brand_stmt = select(Brand).where(
                Brand.tenant_id == tenant_id_int, Brand.name == brand_str
            )
            brand_result = await db.execute(brand_stmt)
            brand = brand_result.scalars().first()

            if not brand:
                try:
                    brand = Brand(name=brand_str, tenant_id=tenant_id_int)
                    db.add(brand)
                    await db.flush()
                except Exception:
                    await db.rollback()
                    brand_result = await db.execute(brand_stmt)
                    brand = brand_result.scalars().first()

            brand_id = brand.id if brand else None
            # Ensure brand_name string is never None fallback to brand_str
            resolved_brand_name = brand.name if (brand and brand.name) else brand_str

            print("data", tenant_id, product_name, sku, mpn, resolved_brand_name)

            # 3. Check Uniqueness at (tenant_id, name, sku, mpn) level
            existing_product_stmt = select(Product).where(
                Product.tenant_id == tenant_id_int,
                Product.name == product_name,
                Product.sku == sku,
                Product.mpn == mpn,
                Product.brand_name == resolved_brand_name,
            )
            existing_result = await db.execute(existing_product_stmt)
            existing_product = existing_result.scalars().first()

            if existing_product:
                print(
                    f"[SKIP] Product already exists for tenant {tenant_id_int}: {product_name} (SKU: {sku})",
                    flush=True,
                )
                return

            # 4. Create and Save the New Product
            new_product = Product(
                tenant_id=tenant_id_int,
                brand_id=brand_id,
                brand=brand,
                brand_name=resolved_brand_name,
                name=product_name,
                sku=sku,
                mpn=mpn,
                product_url=product_url,
                created_by=user_id_int,
            )

            db.add(new_product)
            await db.commit()
            print(f"[SUCCESS] Product created: {product_name}", flush=True)

        except Exception as e:
            await db.rollback()
            print(
                f"[ERROR] Failed to process row for tenant {tenant_id}: {str(e)}",
                flush=True,
            )
            raise


@router.post("/bulk-upload/")
async def upload_excel_data(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    user: dict = Depends(validate_jwt_token),
):
    try:
        contents = await file.read()
        excel_file = BytesIO(contents)

        workbook = load_workbook(excel_file)
        sheet = workbook.active
        data = list(sheet.iter_rows(values_only=True))

        if not data:
            raise HTTPException(status_code=400, detail="Excel file is empty")

        # Clean header strings
        headers = [str(h).strip().lower() if h else "" for h in data[0]]
        rows = data[1:]

        validation = await validate_headers(headers, PRODUCT_TEMPLATE_HEADERS)
        if validation:
            raise HTTPException(
                status_code=400,
                detail=f"Missing Columns: {', '.join(validation)}. Kindly use the explicit corporate template file.",
            )

        tenant_id = user.get("tenant_id")
        if not tenant_id:
            raise HTTPException(
                status_code=400,
                detail="Project Id is required, Please Log in again",
            )

        user_id = user.get("id") or user.get("user_id")
        session_factory = SessionLocal

        task_count = 0

        for row in rows:
            if not any(row):
                continue

            row_dict = dict(zip(headers, row))

            print("row dict", row_dict)

            cleaned_payload = {
                "product_name": (
                    str(row_dict["product name"]).strip()
                    if row_dict.get("product name")
                    else None
                ),
                "brand_name": (
                    str(row_dict["brand name"]).strip()
                    if row_dict.get("brand name")
                    else None
                ),
                "sku": (str(row_dict["sku"]).strip() if row_dict.get("sku") else None),
                "mpn": (str(row_dict["mpn"]).strip() if row_dict.get("mpn") else None),
                "product_url": (
                    str(row_dict["product url"]).strip()
                    if row_dict.get("product url")
                    else None
                ),
            }
            print("data", cleaned_payload)
            background_tasks.add_task(
                process_product_row,
                row_data=cleaned_payload,
                tenant_id=tenant_id,
                user_id=user_id,
                session_factory=session_factory,
            )
            task_count += 1

        return {
            "status": "success",
            "message": f"Bulk processing initiated for {task_count} rows.",
            "task_count": task_count,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Bulk process initialization failed: {str(e)}"
        )
