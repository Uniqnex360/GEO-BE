from io import BytesIO
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse

from app.core.database import get_db, SessionLocal
from app.core.config import settings
from app.core.security import validate_jwt_token
from app.services import ChatService, run_geo_audit_stream
from app.helpers import ExcelTemplateBulider, validate_headers

router = APIRouter()


@router.post("/init_llm_analyzes/")
async def init_llm_analyzes(
    body: dict,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(validate_jwt_token),
):
    """
    example body
    {
    'product_name': 'test',
    'product_url': 'https://www.chmarine.com/international-cruiser-250-antifoul-3L/',
    'extra_context': 'test',
    'model': 'gpt-5-nano'}
    """

    return StreamingResponse(
        ChatService(openai_api_key=settings.OPENAI_API_KEY).start_analysis(
            db, body, user.get("tenant_id")
        ),
        media_type="application/x-ndjson",
    )


CHAT_TEMPLATE_HEADERS = [
    {
        "id": "product_name",
        "identity": "Product Name",
        "required": False,
        "comment": "Name of the Product",
    },
    {
        "id": "product_url",
        "identity": "Product URL",
        "required": False,
        "comment": "Product landing page URL",
    },
    {
        "id": "website",
        "identity": "Website",
        "required": True,
        "comment": "Brand or company website/domain",
    },
    {
        "id": "sku",
        "identity": "SKU",
        "required": False,
        "comment": "Stock Keeping Unit",
    },
    {
        "id": "mpn",
        "identity": "MPN",
        "required": False,
        "comment": "Manufacturer Part Number",
    },
    {
        "id": "ean",
        "identity": "EAN",
        "required": False,
        "comment": "European Article Number (optional)",
    },
    {
        "id": "upc",
        "identity": "UPC",
        "required": False,
        "comment": "Universal Product Code",
    },
    {
        "id": "country",
        "identity": "Country",
        "required": False,
        "comment": "Target country for the GEO audit",
    },
    {
        "id": "extra_context",
        "identity": "Extra Context",
        "required": False,
        "comment": "Additional instructions or context",
    },
    {
        "id": "model_choice",
        "identity": "Model Choice",
        "required": False,
        "comment": "LLM model (e.g. GPT, Gemini, Claude)",
    },
]

CHAT_EXAMPLE_DATA = [
    {
        "product_name": "Premium Brake Pads",
        "product_url": "https://example.com/products/premium-brake-pads",
        "website": "https://stopsafe.com",
        "sku": "SKU-10001",
        "mpn": "BP-9923-X",
        "ean": "4006381333931",
        "upc": "884616012345",
        "country": "United States",
        "extra_context": "Focus on automotive aftermarket competitors.",
        "model_choice": "GPT",
    },
]


@router.get("/bulk-upload-template/")
async def generate_chat_template(
    user: dict = Depends(validate_jwt_token),
):
    """api endpoint that streams bulk upload template for chat"""

    builder = ExcelTemplateBulider(
        headers=CHAT_TEMPLATE_HEADERS,
        sheet_name="chat_template",
        data=CHAT_EXAMPLE_DATA,
        example=True,
    )

    wb = builder.build()

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachement; filename=industry_template.xlsx"},
    )


from pydantic import BaseModel, Field
from typing import Optional
from app.models.base import LLMModels


class GEOAuditRequest(BaseModel):
    """V2 Flexible Request Inputs for multiple source identification types."""

    product_name: Optional[str] = Field(None, description="Name of the target product")
    product_url: Optional[str] = Field(
        None, description="Target product landing page URL"
    )
    website: Optional[str] = Field(None, description="Brand/corporate target domain")
    sku: Optional[str] = Field(None, description="Stock Keeping Unit number")
    mpn: Optional[str] = Field(None, description="Manufacturer Part Number")
    upc: Optional[str] = Field(None, description="Universal Product Code")
    country: Optional[str] = Field(None, description="Target geographical focus region")
    extra_context: Optional[str] = Field(
        None, description="Additional context parameter text"
    )
    model_choice: LLMModels = Field(
        default=LLMModels.GPT, description="Selected LLM execution engine"
    )


async def process_row_background_task(
    row_data: dict, tenant_id: int, user_id: int, session_factory
):
    """Worker function acting as the isolated background task boundary per row.

    Since generators ('yield') cannot be awaited directly inside background tasks,
    this driver consumes the streaming outputs and handles errors safely per task execution context.
    """
    # Create an isolated database session block per row context loop to avoid session contention
    async with session_factory() as db:
        try:
            # Consume generator results safely without breaking execution loops
            # async for log in chat_service.start_analysis(
            #     db=db, data=row_data, tenant_id=tenant_id, user_id=user_id
            # ):
            async for log in run_geo_audit_stream(
                payload=row_data,
                db=db,
                tenant_id=tenant_id,
                user_id=user_id,
            ):
                # Optionally stream or log pipeline events to external cloud monitoring engines
                pass
        except Exception as task_err:
            print(
                f"[Row Worker Error] Failed processing for row identification parameters {row_data}: {str(task_err)}"
            )


@router.post("/bulk-upload/")
async def upload_excel(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    user: dict = Depends(validate_jwt_token),
):
    try:
        contents = await file.read()
        excel_file = BytesIO(contents)

        workbook = load_workbook(excel_file)
        sheet = workbook.active
        data = list(sheet.iter_rows(values_only=True))

        if not data:
            raise HTTPException(status_code=400, detail="Excel file is empty")

        # Map header string inputs down to cleaner tracking coordinates
        headers = [str(h).strip().lower() if h else "" for h in data[0]]
        rows = data[1:]

        validation = await validate_headers(headers, CHAT_TEMPLATE_HEADERS)

        if validation:
            raise HTTPException(
                status_code=400,
                detail=f"Missing Columns: {', '.join(validation)}. Kindly use the explicit corporate template file.",
            )

        # Extraction parameters context
        tenant_id = user.get("tenant_id", 1)
        user_id = user.get("user_id")

        # Session factory tracker used to generate clean isolated context wrappers inside background loops
        # This prevents session collisions across concurrently executing background workers
        session_factory = SessionLocal

        task_count = 0

        # -------- Parse Matrix Rows Into Background Queue Payload --------
        for row in rows:
            if not any(row):  # Skip completely empty rows
                continue

            # Zip explicit header labels cleanly alongside values matrix entries
            row_dict = dict(zip(headers, row))

            # Clean blank string parameters or text inputs safely
            cleaned_payload = {
                "product_name": (
                    str(row_dict["product_name"]).strip()
                    if row_dict.get("product_name")
                    else None
                ),
                "product_url": (
                    str(row_dict["product_url"]).strip()
                    if row_dict.get("product_url")
                    else None
                ),
                "website": (
                    str(row_dict["website"]).strip()
                    if row_dict.get("website")
                    else None
                ),
                "sku": (str(row_dict["sku"]).strip() if row_dict.get("sku") else None),
                "mpn": (str(row_dict["mpn"]).strip() if row_dict.get("mpn") else None),
                "upc": (str(row_dict["upc"]).strip() if row_dict.get("upc") else None),
                "country": (
                    str(row_dict["country"]).strip()
                    if row_dict.get("country")
                    else "United States of America"
                ),
                "extra_context": (
                    str(row_dict["extra_context"]).strip()
                    if row_dict.get("extra_context")
                    else ""
                ),
                "model_choice": (
                    LLMModels(row_dict["model"])
                    if row_dict.get("model")
                    else LLMModels.GPT
                ),
            }
            # Append the processing run straight to background execution queue loops
            background_tasks.add_task(
                process_row_background_task,
                row_data=cleaned_payload,
                tenant_id=tenant_id,
                user_id=user_id,
                session_factory=session_factory,
            )
            task_count += 1

        return {
            "status": "Accepted",
            "message": f"Successfully scheduled {task_count} product validation workflows in the background.",
            "processed_rows": task_count,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Bulk process initialization failed: {str(e)}"
        )
